# =============================================================================
# RE_VE_PROCESSOR
# =============================================================================
# Purpose:
#   Calculates the theoretical nozzle exit velocity V_noz and associated
#   Reynolds numbers (Re_noz and Re_cyl) from the wind-tunnel differential
#   pressure measurement and the experiment geometry. The calculation chain
#   applies Bernoulli's equation to obtain the centerline tunnel velocity from
#   the differential pressure, derives air density and dynamic viscosity via
#   the ideal gas law and Sutherland's law, and then applies a flow-profile
#   correction factor C_profile (computed automatically from the tunnel
#   Reynolds regime: 0.50 for laminar, power-law interpolation for turbulent)
#   together with the tunnel-to-nozzle area ratio to obtain V_noz. Re_noz is
#   based on the nozzle slot height B, and Re_cyl on the cylinder diameter D.
#   Each set of inputs is logged to a history table. Results can be exported
#   as a single-case Flow_Data.csv/xlsx (specific mode) or as a cumulative
#   synthesis file (global mode) for multi-case parameter tracking.
#
# Inputs:
#   - User-supplied GUI inputs: differential pressure dP (Pa), air temperature
#     T (deg C), atmospheric pressure P_atm (Pa), tunnel diameter D1 (mm),
#     nozzle width W_noz (mm), nozzle height B (mm), cylinder diameter
#     D2 (mm)
#
# Outputs:
#   - "Specific" export mode (one case):
#       experiments/<case>/Processing_Parameters/Flow_Data.csv / .xlsx
#   - "Global" export mode (cumulative history):
#       <workspace>/Flow_Data_Synthese.csv / .xlsx
#
# Dependencies:
#   - Imported by Re_comparison_processor.py (calculate_physics method)
#   - Imported by pipeline.py (Step 1) and usage.py (hub launcher)
#
# Usage:
#   - Standalone: python Re_Ve_processor.py  (opens Tkinter GUI)
#   - Via pipeline: called programmatically by pipeline.py (Step 1)
#   - Via hub:      launched from usage.py as "Re and Velocity Calculator"
# =============================================================================

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import math
import csv
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os
import sys as _sys
import pandas as _pd

# --- CONFIG LOADER ---
def _load_cfg(filename):
    # All configs are .xlsx (callers may still pass the legacy .csv name).
    filename = filename.replace('.csv', '.xlsx')
    _p = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'config', filename)
    _df = _pd.read_excel(_p, index_col='parameter')
    return _df['value'] if 'value' in _df.columns else _df.iloc[:, 0]

try:
    _cfg_geom  = _load_cfg('config_geometry.csv')
    _cfg_flow  = _load_cfg('config_flow_regime.csv')
    _RE_LAM = float(_cfg_flow['re_laminar_upper'])
    _RE_TR  = float(_cfg_flow['re_transition_upper'])
    _C_LAM  = float(_cfg_flow['c_profile_laminar'])
    _N_COEF = float(_cfg_flow['c_profile_turbulent_log_coeff'])
    _N_OFF  = float(_cfg_flow['c_profile_turbulent_log_offset'])
except Exception:
    _cfg_geom  = None
    _cfg_flow  = None
    _RE_LAM = 2300.0
    _RE_TR  = 4000.0
    _C_LAM  = 0.50
    _N_COEF = 1.22
    _N_OFF  = 1.0

# Channel -> pitot velocity calibration (affine): V_pitot = a*V_channel + b.
# Maps the channel-derived nozzle velocity onto the pitot metrological reference
# (the channel typically reads systematically low vs the pitot). Constants from
# config/config_calibration.xlsx; fallback identity (a=1, b=0).
try:
    _cal = _pd.read_excel(
        os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..',
                     'config', 'config_calibration.xlsx')).set_index('parameter')['value']
    _CH_A = float(_cal['channel_slope'])
    _CH_B = float(_cal['channel_intercept'])
except Exception:
    _CH_A, _CH_B = 1.0, 0.0

# Import shared air properties utility
_sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'utils'))
# Shared RSS uncertainty-propagation helpers (Task 3). All uncertainty values
# (incl. the channel fit SE and per-dimension geometry) come from
# config_uncertainty.xlsx via the uncertainty module.
try:
    from uncertainty import (rel_area_ratio as _u_rel_area,
                             rel_velocity_channel as _u_rel_vchan,
                             apply_calibration_u as _u_apply_cal,
                             U_DP_PA as _U_DP, CHANNEL_FIT_SE as _CH_SE,
                             U_D1_MM as _U_D1, U_W_MM as _U_W,
                             U_B_MM as _U_B, U_D2_MM as _U_D2)
    _HAS_UNC = True
except Exception:
    _HAS_UNC = False
    _U_DP = 1.0
    _CH_SE = 0.0   # no calibration fit residual by default (identity map)
    _U_D1, _U_W, _U_B, _U_D2 = 2.0, 0.5, 0.1, 0.1
try:
    from air_properties import calc_air_properties, R_AIR, MU_REF, T_SUT, S_SUT, T_FALLBACK, P_FALLBACK
except Exception:
    # Fallback constants if air_properties module is unavailable
    R_AIR = 287.05; MU_REF = 1.716e-5; T_SUT = 273.15; S_SUT = 110.4
    T_FALLBACK = 20.0; P_FALLBACK = 101325.0
    def calc_air_properties(temp_c, p_atm):
        T = temp_c + 273.15
        rho = p_atm / (R_AIR * T)
        mu  = MU_REF * (T / T_SUT)**1.5 * (T_SUT + S_SUT) / (T + S_SUT)
        k   = 0.0242 + 0.00007 * temp_c
        Pr  = (mu * 1005.0) / k
        return rho, mu, k, Pr

def compute_flow_uncertainties(dp_pa, v_tun_avg, v_noz, v_noz_channel,
                               re_tun, re_noz, re_cyl,
                               d1_mm, w2_mm, h2_mm, d2_mm, c_profile=None):
    """1-sigma absolute uncertainties for the Re/velocity chain (see
    ../../docs/UNCERTAINTY_SPEC.md §3). Velocity uncertainty propagates the channel
    measurement (dP sqrt-law + area ratio) through the channel->pitot affine
    calibration:  u(V_noz) = sqrt[(a_ch*u_channel)^2 + SE_ch^2]; Reynolds numbers
    inherit the relative velocity uncertainty (rho, mu, L are second-order).
    Returns NaN for every field if the uncertainty helper or a non-positive input
    makes it ill-defined."""
    nan = float('nan')
    keys = ['u_V_tun_ms', 'u_V_noz_ms', 'u_V_noz_channel_raw_ms',
            'u_Re_tun', 'u_Re_noz', 'u_Re_cyl']
    if not _HAS_UNC or dp_pa <= 0 or v_noz <= 0:
        return {k: nan for k in keys}

    rel_area   = _u_rel_area(d1_mm, w2_mm, h2_mm)            # uR/R ~ 2.05 % (V_noz; per-dim)
    rel_vtun   = 0.5 * _U_DP / dp_pa                          # Bernoulli sqrt-law
    rel_vchan  = _u_rel_vchan(dp_pa, rel_area)               # sqrt[(1/2 udP/dP)^2 + (uR/R)^2]
    # (No power-law / C_profile term: C_profile is fixed at 1.0 — the empirical
    # channel->pitot calibration subsumes the profile. See calculate_physics.)
    u_vchan    = v_noz_channel * rel_vchan
    u_v_noz    = _u_apply_cal(u_vchan, _CH_A, _CH_SE)        # affine calibration rule
    rel_v_noz  = u_v_noz / v_noz

    # Reynolds CHARACTERISTIC-length terms (the same measured lengths reappear as
    # the Re length scale, separate from their area-ratio role in V_noz). Per-dim
    # 1-sigma from config_uncertainty.xlsx (_U_D1/_U_W/_U_B/_U_D2 mm):
    rel_B  = _U_B  / h2_mm if h2_mm > 0 else 0.0
    rel_D  = _U_D2 / d2_mm if d2_mm > 0 else 0.0
    rel_D1 = _U_D1 / d1_mm if d1_mm > 0 else 0.0

    # Re_noz (length B): B CANCELS the area-ratio B (V_noz ∝ 1/B, Re ∝ V_noz·B ⇒
    #   B-independent), so the area-ratio uB/B is removed (affine residual ~0.2 %).
    rel_re_noz = max(0.0, rel_v_noz**2 - rel_B**2) ** 0.5
    # Re_cyl (length D): D is independent of V_noz, so uD/D adds in quadrature.
    rel_re_cyl = (rel_v_noz**2 + rel_D**2) ** 0.5
    # Re_tun (length D1; v_tun has no area ratio): pressure ⊕ uD1/D1.
    rel_re_tun = (rel_vtun**2 + rel_D1**2) ** 0.5

    return {
        'u_V_tun_ms':              v_tun_avg * rel_vtun,
        'u_V_noz_ms':              u_v_noz,
        'u_V_noz_channel_raw_ms':  u_vchan,
        'u_Re_tun':                re_tun * rel_re_tun,
        'u_Re_noz':                re_noz * rel_re_noz,
        'u_Re_cyl':                re_cyl * rel_re_cyl,
    }


class ReVeCalculatorApp:
    def __init__(self, root, default_folder=None):
        self.root = root
        self.root.title("Tunnel and Nozzle Flow Dynamics")
        self.root.geometry("1200x950")
        self.history = []
        self.default_folder = default_folder
        
        # --- Input Variables ---
        self.delta_p_var = tk.StringVar(value="12.0")
        self.temp_var = tk.StringVar(value="21.0")
        self.p_atm_var = tk.StringVar(value="100050")
        
        # Geometry defaults (loaded from config where appropriate)
        self.d1_var = tk.StringVar(value=str(float(_cfg_geom['tunnel_diameter_D1'])) if _cfg_geom is not None else "200.0")
        self.w2_var = tk.StringVar(value=str(float(_cfg_geom['nozzle_width_cyl']))   if _cfg_geom is not None else "150.0")
        self.h2_var = tk.StringVar(value=str(float(_cfg_geom['nozzle_height_B']))    if _cfg_geom is not None else "18.0")
        self.d2_var = tk.StringVar(value="6.52")  # Per-session input; varies by experiment

        # Export Mode
        self.export_mode_var = tk.StringVar(value="specific")

        self.setup_ui()

    def calculate_physics(self, dp_pa, temp_c, p_atm, d1_mm, w2_mm, h2_mm, d2_mm):
        # 1. Propriétés du fluide (via shared air_properties utility)
        rho, mu, _, _ = calc_air_properties(temp_c, p_atm)

        # 2. Géométrie
        d1 = d1_mm / 1000.0
        h2 = h2_mm / 1000.0
        d2 = d2_mm / 1000.0
        a_tun = math.pi * (d1 / 2.0)**2
        a_noz = (w2_mm * h2_mm) / 1e6
        k_ratio = a_tun / a_noz

        # 3. Vitesse absolue au centre (mesurée par la sonde)
        v_tun_center = math.sqrt((2 * dp_pa) / rho) if dp_pa > 0 else 0.0

        # 4. Calcul du Reynolds max (au centre) pour connaitre le régime
        re_max = (rho * v_tun_center * d1) / mu

        # 5. Profile factor C_profile — DISABLED (fixed at 1.0).
        # The power-law centerline->mean correction is intentionally NOT applied.
        # The channel->pitot calibration is EMPIRICAL and was fit directly on the
        # raw continuity estimate sqrt(2dP/rho)*(A_ch/A_noz) with NO profile factor,
        # so it already absorbs the centerline-vs-mean profile (and every other
        # systematic). Applying C_profile here would double-correct the same effect
        # and be inconsistent with the basis on which the calibration was derived.
        # (The old regime-dependent power law lives in git history / README.)
        c_profile = 1.0

        # 6. Vitesses et Reynolds finaux
        v_tun_avg = v_tun_center * c_profile
        v_noz_channel = v_tun_avg * k_ratio          # raw channel-derived (traceability only)
        # THE nozzle velocity is expressed in the PITOT reference via the
        # channel->pitot affine calibration (config_calibration.xlsx):
        #     V_noz = a*V_channel + b
        # so v_noz below — and every Reynolds number derived from it — is in the
        # pitot reference used by the Cobra.
        v_noz = _CH_A * v_noz_channel + _CH_B

        # NOTE: the idealised gap (throat) velocity V_gap = V_noz*B/(B-D) has been
        # removed. It assumed full closed-channel confinement, but the cylinder sits
        # at an OPEN jet exit so it over-estimated the real throat speed; the
        # Strouhal number now uses the MEASURED peak velocity U_max, and the velocity
        # normalisation (U/U0) uses V_noz for both free and cylinder cases.
        re_tun = (rho * v_tun_avg * d1) / mu
        re_noz = (rho * v_noz * h2) / mu
        re_cyl = (rho * v_noz * d2) / mu

        return (v_tun_avg, v_noz, re_tun, re_noz, re_cyl, rho, mu, k_ratio,
                c_profile, v_noz_channel)

    def on_calculate(self):
        try:
            dp = float(self.delta_p_var.get())
            t = float(self.temp_var.get())
            patm = float(self.p_atm_var.get())
            
            d1 = float(self.d1_var.get())
            w2 = float(self.w2_var.get())
            h2 = float(self.h2_var.get())
            d2 = float(self.d2_var.get())
            
            (v_tun, v_noz, re_tun, re_noz, re_cyl, rho, mu, k, c_prof_dyn,
             v_noz_channel) = self.calculate_physics(dp, t, patm, d1, w2, h2, d2)

            self.res_v_noz.config(text=f"{v_noz:.2f} m/s", fg="#007acc")
            self.res_re_noz.config(text=f"{re_noz:.0f}", fg="#d9534f")
            self.res_re_cyl.config(text=f"{re_cyl:.0f}", fg="#d9534f") 
            self.res_rho.config(text=f"{rho:.3f} kg/m³")
            self.res_k.config(text=f"{k:.2f}")
            self.res_cprof.config(text=f"{c_prof_dyn:.3f}")

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # 1-sigma uncertainties (UNCERTAINTY_SPEC.md §3); stored as the column
            # immediately after each value (value in N, u_ in N+1).
            u = compute_flow_uncertainties(dp, v_tun, v_noz, v_noz_channel,
                                           re_tun, re_noz, re_cyl,
                                           d1, w2, h2, d2, c_profile=c_prof_dyn)
            entry = {
                "Time": timestamp,
                "dP_Pa": dp,
                "Temp_C": t,
                "P_atm_Pa": patm,
                "C_prof": f"{c_prof_dyn:.3f}",
                "D1_mm": d1,
                "D2_mm": d2,
                "W_noz_mm": w2,
                "B_mm": h2,
                "Rho_kgm3": f"{rho:.3f}",
                "K_ratio": f"{k:.2f}",
                "V_tun_ms": f"{v_tun:.2f}",
                "u_V_tun_ms": f"{u['u_V_tun_ms']:.3f}",
                "V_noz_ms": f"{v_noz:.2f}",
                "u_V_noz_ms": f"{u['u_V_noz_ms']:.3f}",
                "V_noz_channel_raw_ms": f"{v_noz_channel:.2f}",
                "u_V_noz_channel_raw_ms": f"{u['u_V_noz_channel_raw_ms']:.3f}",
                "Re_tun": int(re_tun),
                "u_Re_tun": f"{u['u_Re_tun']:.0f}",
                "Re_noz": int(re_noz),
                "u_Re_noz": f"{u['u_Re_noz']:.0f}",
                "Re_cyl": int(re_cyl),
                "u_Re_cyl": f"{u['u_Re_cyl']:.0f}",
                "Internal_ID": timestamp + str(len(self.history))
            }
            self.history.append(entry)
            self.update_history_display()

        except ValueError:
            messagebox.showerror("Input Error", "Please verify all numeric inputs.")

    def update_history_display(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        for r in reversed(self.history):
            self.tree.insert("", "end", iid=r["Internal_ID"], values=(
                r["Time"], r["dP_Pa"], r["Temp_C"], r["P_atm_Pa"], r["C_prof"],
                r["D1_mm"], r["D2_mm"], r["W_noz_mm"], r["B_mm"],
                r["Rho_kgm3"], r["K_ratio"], r["V_tun_ms"], r["V_noz_ms"], r["Re_tun"], r["Re_noz"], r["Re_cyl"]
            ))

    def delete_selected(self):
        selected_item = self.tree.selection()
        if not selected_item: return
        self.history = [h for h in self.history if h["Internal_ID"] != selected_item[0]]
        self.tree.delete(selected_item)

    def export_data(self):
        if not self.history:
            messagebox.showwarning("Export Warning", "History is empty. Nothing to export.")
            return

        # Define the base workspace directory
        base_folder = self.default_folder
        if not base_folder or not os.path.exists(base_folder):
            base_folder = filedialog.askdirectory(title="Sélectionner le dossier cible")
            if not base_folder:
                return

        export_mode = self.export_mode_var.get()
        header = [k for k in self.history[0].keys() if k != "Internal_ID"]

        try:
            if export_mode == "specific":
                # Mode Cas Spécifique : 1 ligne, dans /Processing_Parameters/
                target_dir = os.path.join(base_folder, "Processing_Parameters")
                os.makedirs(target_dir, exist_ok=True)
                
                csv_path = os.path.join(target_dir, "Flow_Data.csv")
                xlsx_path = os.path.join(target_dir, "Flow_Data.xlsx")
                
                data_to_export = [self.history[-1]] # Uniquement la dernière ligne calculée
                
                # Écrasement CSV (mode 'w')
                with open(csv_path, 'w', newline='') as f:
                    w = csv.DictWriter(f, fieldnames=header, extrasaction='ignore')
                    w.writeheader()
                    w.writerows(data_to_export)
                    
                # Écrasement Excel
                wb = Workbook()
                ws = wb.active
                ws.title = "Flow Data"
                ws.append(header)
                row = [data_to_export[0].get(h, "") for h in header]
                ws.append(row)
                wb.save(xlsx_path)

            else:
                # Mode Synthèse Globale : Toutes les lignes, à la racine
                csv_path = os.path.join(base_folder, "Flow_Data_Synthèse.csv")
                xlsx_path = os.path.join(base_folder, "Flow_Data_Synthèse.xlsx")
                
                data_to_export = self.history
                
                # Ajout ou Création CSV (mode 'a')
                file_exists = os.path.isfile(csv_path)
                with open(csv_path, 'a' if file_exists else 'w', newline='') as f:
                    w = csv.DictWriter(f, fieldnames=header, extrasaction='ignore')
                    if not file_exists:
                        w.writeheader()
                    w.writerows(data_to_export)
                    
                # Ajout ou Création Excel
                if os.path.isfile(xlsx_path):
                    wb = openpyxl.load_workbook(xlsx_path)
                    ws = wb.active
                    for entry in data_to_export:
                        row = [entry.get(h, "") for h in header]
                        ws.append(row)
                else:
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Flow Data Synthesis"
                    ws.append(header)
                    for entry in data_to_export:
                        row = [entry.get(h, "") for h in header]
                        ws.append(row)
                wb.save(xlsx_path)

            messagebox.showinfo("Success", f"Exportation réussie ({'Cas Spécifique' if export_mode == 'specific' else 'Synthèse Globale'}) :\n\n{csv_path}\n{xlsx_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"An error occurred: {e}")
    
    def setup_ui(self):
        p_frame = tk.LabelFrame(self.root, text=" 1. Inputs (Environment & Metrology) ", padx=10, pady=10)
        p_frame.pack(padx=20, pady=10, fill="x")
        
        tk.Label(p_frame, text="Delta P Tunnel (Pa):").grid(row=0, column=0, sticky="w")
        tk.Entry(p_frame, textvariable=self.delta_p_var, width=12).grid(row=0, column=1)
        tk.Label(p_frame, text="Temp (°C):").grid(row=0, column=2, padx=10, sticky="w")
        tk.Entry(p_frame, textvariable=self.temp_var, width=12).grid(row=0, column=3)
        tk.Label(p_frame, text="P_atm (Pa):").grid(row=0, column=4, padx=10, sticky="w")
        tk.Entry(p_frame, textvariable=self.p_atm_var, width=12).grid(row=0, column=5)

        g_frame = tk.LabelFrame(self.root, text=" 2. Geometry ", padx=10, pady=10)
        g_frame.pack(padx=20, pady=5, fill="x")
        
        tk.Label(g_frame, text="D1 Tunnel (mm):").grid(row=0, column=0)
        tk.Entry(g_frame, textvariable=self.d1_var, width=10).grid(row=0, column=1)
        tk.Label(g_frame, text="W_noz (mm):").grid(row=0, column=2, padx=10)
        tk.Entry(g_frame, textvariable=self.w2_var, width=10).grid(row=0, column=3)
        tk.Label(g_frame, text="B (mm):").grid(row=0, column=4, padx=10)
        tk.Entry(g_frame, textvariable=self.h2_var, width=10).grid(row=0, column=5)
        tk.Label(g_frame, text="D2 Cyl (mm):").grid(row=0, column=6, padx=10)
        tk.Entry(g_frame, textvariable=self.d2_var, width=10).grid(row=0, column=7)

        r_frame = tk.LabelFrame(self.root, text=" 3. Main Results (Nozzle Exit) ", padx=10, pady=10)
        r_frame.pack(padx=20, pady=10, fill="x")
        
        tk.Label(r_frame, text="V_noz:", font=("Arial", 12)).grid(row=0, column=0)
        self.res_v_noz = tk.Label(r_frame, text="0.00", font=("Arial", 14, "bold"))
        self.res_v_noz.grid(row=0, column=1, padx=20)
        
        tk.Label(r_frame, text="Re_noz:", font=("Arial", 12)).grid(row=0, column=2)
        self.res_re_noz = tk.Label(r_frame, text="0", font=("Arial", 14, "bold"))
        self.res_re_noz.grid(row=0, column=3, padx=20)
        
        tk.Label(r_frame, text="Auto C_prof:").grid(row=0, column=4, padx=5)
        self.res_cprof = tk.Label(r_frame, text="0.000", fg="green", font=("Arial", 11, "bold"))
        self.res_cprof.grid(row=0, column=5, padx=5)

        tk.Label(r_frame, text="Rho:").grid(row=0, column=6, padx=5)
        self.res_rho = tk.Label(r_frame, text="1.204", fg="blue")
        self.res_rho.grid(row=0, column=7, padx=5)
        
        tk.Label(r_frame, text="Area k:").grid(row=0, column=8, padx=5)
        self.res_k = tk.Label(r_frame, text="1.0")
        self.res_k.grid(row=0, column=9)

        tk.Label(r_frame, text="Re_cyl:", font=("Arial", 12)).grid(row=1, column=2, pady=5)
        self.res_re_cyl = tk.Label(r_frame, text="0", font=("Arial", 14, "bold"))
        self.res_re_cyl.grid(row=1, column=3, padx=20, pady=5)

        # --- Controls Area ---
        controls_frame = tk.Frame(self.root)
        controls_frame.pack(pady=10, fill="x", padx=20)
        
        calc_del_frame = tk.Frame(controls_frame)
        calc_del_frame.pack(side="left", expand=True, fill="x")
        
        tk.Button(calc_del_frame, text="CALCULATE", command=self.on_calculate, bg="#28a745", fg="white", font=("Arial", 11, "bold"), height=2).pack(side="left", expand=True, fill="x", padx=5)
        tk.Button(calc_del_frame, text="DELETE", command=self.delete_selected, bg="#dc3545", fg="white", font=("Arial", 11, "bold"), height=2).pack(side="left", expand=True, fill="x", padx=5)
        
        export_options_frame = tk.LabelFrame(controls_frame, text=" Options d'Exportation ")
        export_options_frame.pack(side="left", expand=True, fill="x", padx=5)
        
        tk.Radiobutton(export_options_frame, text="Cas Spécifique (1 ligne dans Processing_Parameters)", variable=self.export_mode_var, value="specific").pack(anchor="w", padx=10)
        tk.Radiobutton(export_options_frame, text="Synthèse Globale (Historique dans le dossier source)", variable=self.export_mode_var, value="global").pack(anchor="w", padx=10)

        tk.Button(controls_frame, text="EXPORT CSV + EXCEL", command=self.export_data, bg="#6c757d", fg="white", font=("Arial", 11, "bold"), height=2).pack(side="left", expand=True, fill="x", padx=5)

        l_frame = tk.LabelFrame(self.root, text=" Complete History Log ", padx=10, pady=10)
        l_frame.pack(padx=20, pady=10, fill="both", expand=True)
        
        columns = ("Time", "dP", "T", "P_atm", "C_prof", "D1", "D2", "W_noz", "B_mm", "Rho", "k", "V_tun", "V_noz", "Re_tun", "Re_noz", "Re_cyl")
        self.tree = ttk.Treeview(l_frame, columns=columns, show="headings")
        
        for col in columns:
            self.tree.heading(col, text=col)
            if "Re" in col or "V" in col: width = 80
            elif "noz" in col or "D1" in col or "D2" in col or "B_mm" in col: width = 60
            elif "C_prof" in col: width = 65
            else: width = 65
            self.tree.column(col, width=width, anchor="center")
            
        self.tree.pack(side="left", fill="both", expand=True)
        sb = ttk.Scrollbar(l_frame, orient="vertical", command=self.tree.yview); self.tree.configure(yscrollcommand=sb.set); sb.pack(side="right", fill="y")

if __name__ == "__main__":
    root = tk.Tk()
    app = ReVeCalculatorApp(root)
    root.mainloop()