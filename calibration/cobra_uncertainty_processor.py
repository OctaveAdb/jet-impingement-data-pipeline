# =============================================================================
# COBRA_UNCERTAINTY_PROCESSOR  (no GUI)
# =============================================================================
# Reads each TFI Cobra .thA time-history file (65,536 samples of u, v, w),
# computes the mean flow speed |V| = sqrt(u^2+v^2+w^2) and propagates the
# measurement uncertainty to the MEAN.
#
# Uncertainty model for the mean
# -------------------------------
# The flow oscillates at LOW frequency, so the 65,536 samples are strongly
# autocorrelated -> the effective number of independent samples is far smaller
# than N and the naive s/sqrt(N) underestimate is not physical. The mean itself
# is well converged (several oscillation periods are captured), so its scatter
# is not the limiting factor. The limiting factor is the probe's absolute
# accuracy, which is a systematic floor that does NOT average down:
#
#   u_mean = delta = 0.5 m/s        (probe absolute accuracy, fixed floor)
#
# (s/sqrt(N) is still computed and printed for reference only.)
#
# Output: writes Mean speed + uncertainty into Calibration.xlsx and prints a
# table. Standalone:  python cobra_uncertainty_processor.py
# =============================================================================

import os, glob, re
import numpy as np

DELTA = 0.5          # cobra per-sample absolute resolution [m/s]
HERE  = os.path.dirname(os.path.abspath(__file__))
THA_DIR = os.path.join(HERE, "Velocity Calibration_2")
XLSX    = os.path.join(HERE, "Calibration.xlsx")


# --- binary .thA reader (kept from read_th_file_processor, trimmed) ----------
def _read_header_strings(fid):
    np.fromfile(fid, dtype=np.int16, count=4)
    np.fromfile(fid, dtype=np.int16, count=4)


def read_th_file(file_path):
    with open(file_path, 'rb') as fid:
        fmt = np.fromfile(fid, dtype=np.int32, count=1)[0]
        meta = {}
        if fmt == 3:
            fid.seek(4, 1)
            np.fromfile(fid, dtype=np.int32, count=1)          # device id
            fid.seek(4, 1)
            _read_header_strings(fid)
            n = int(np.fromfile(fid, dtype=np.int32, count=1)[0])
            b = int(np.fromfile(fid, dtype=np.int32, count=1)[0])
            np.fromfile(fid, dtype=np.float64, count=1)        # data rate
            np.fromfile(fid, dtype=np.float64, count=2)        # p_baro, t_mean
            has_pref = bool(np.fromfile(fid, dtype=np.uint8, count=1)[0])
        else:
            np.fromfile(fid, dtype=np.int32, count=1)
            n = int(np.fromfile(fid, dtype=np.int32, count=1)[0])
            b = int(np.fromfile(fid, dtype=np.int32, count=1)[0])
            np.fromfile(fid, dtype=np.float64, count=1)
            if fmt == 2:
                np.fromfile(fid, dtype=np.float64, count=2)
            has_pref = bool(np.fromfile(fid, dtype=np.uint8, count=1)[0])

        u_all, v_all, w_all = [], [], []
        for _ in range(n // b):
            u_all.append(np.fromfile(fid, dtype=np.float32, count=b))
            v_all.append(np.fromfile(fid, dtype=np.float32, count=b))
            w_all.append(np.fromfile(fid, dtype=np.float32, count=b))
            np.fromfile(fid, dtype=np.float32, count=b)        # static pressure
            if has_pref:
                np.fromfile(fid, dtype=np.float32, count=b)    # reference pressure
        return np.concatenate(u_all), np.concatenate(v_all), np.concatenate(w_all)


# --- per-file mean speed + uncertainty of the mean ---------------------------
def process_file(path, delta=DELTA):
    u, v, w = read_th_file(path)
    speed = np.sqrt(u**2 + v**2 + w**2)
    N = speed.size
    mean = float(speed.mean())
    s = float(speed.std(ddof=1))                # sample standard deviation
    sem = s / np.sqrt(N)                          # reference only (autocorrelated)
    u_mean = float(delta)                         # probe accuracy floor: 0.5 m/s
    return dict(N=N, mean=mean, std=s, sem=sem, u_mean=u_mean)


def main():
    files = glob.glob(os.path.join(THA_DIR, "*.thA"))
    # sort by the leading pressure number in the filename ("1Pa", "2Pa", ...)
    def pa(p):
        m = re.match(r'(\d+)Pa', os.path.basename(p))
        return int(m.group(1)) if m else 1e9
    files.sort(key=pa)

    results = []
    print(f"{'Pa':>3} {'N':>7} {'Mean|V|':>9} {'std':>8} "
          f"{'sem(ref)':>9} {'u_mean':>8}")
    for p in files:
        r = process_file(p)
        r['pa'] = pa(p)
        results.append(r)
        print(f"{r['pa']:>3} {r['N']:>7} {r['mean']:>9.4f} {r['std']:>8.4f} "
              f"{r['sem']:>9.5f} {r['u_mean']:>8.2f}")

    # --- write into Excel ----------------------------------------------------
    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX)
        ws = wb['Tabelle1']
        for r in results:
            row = r['pa'] + 1                     # 1 Pa -> row 2 ... 16 Pa -> 17
            ws.cell(row=row, column=3,  value=round(r['mean'],   4))  # C  Cobra mean
            ws.cell(row=row, column=12, value=round(r['u_mean'], 2))  # L  u(Cobra mean)=0.5
        wb.save(XLSX)
        print(f"\nWritten Mean |V| (col C) and u(mean) (col L) into {os.path.basename(XLSX)}")
    except Exception as e:
        print(f"Excel write skipped: {e}")


if __name__ == "__main__":
    main()
